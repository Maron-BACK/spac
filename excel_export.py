"""
엑셀 출력 모듈
- DB의 현재 상태를 그대로 받아 새 xlsx로 출력 (기존 스팩정리.xlsx와 별도 파일)
- 청산일 임박(6개월 이내) 행은 노란색 배경
- 이율 출처가 'KSFC' (추정)인 셀은 회색으로 표시
"""
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import db
import calc


HEADERS = [
    ("ticker",                   "종목코드",       10),
    ("name",                     "종목명",         18),
    ("payment_date",             "납입기일",       12),
    ("estimated_liquidation_date","청산예상일",     12),
    ("days_until_liquidation",   "청산까지(일)",   12),
    ("offering_price",           "공모가",         10),
    ("current_price",            "현재가",         10),
    ("rate_y1",                  "1년차 이율(%)",  12),
    ("rate_y2",                  "2년차 이율(%)",  12),
    ("rate_y3",                  "3년차 이율(%)",  12),
    ("liquidation_value",        "예상청산금",     12),
    ("total_return_pct",         "총수익률(%)",    11),
    ("annualized_return_pct",    "연환산(%)",      10),
    ("rate_source",              "이율출처",       10),
    ("rate_note",                "출처비고",       28),
    ("last_synced",              "최종확인",       18),
    ("note",                     "메모",           20),
]


YELLOW_FILL = PatternFill("solid", fgColor="FFF59D")   # 청산임박 (6개월 이내)
GREY_FILL   = PatternFill("solid", fgColor="E0E0E0")   # KSFC 추정
RED_FILL    = PatternFill("solid", fgColor="FFCDD2")   # 관리종목 임박
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_to_xlsx(out_path: str | Path) -> str:
    """현재 DB 상태를 엑셀로 저장. out_path 반환."""
    out_path = Path(out_path)
    wb = openpyxl.Workbook()

    # ---------- 1. 메인 시트 ----------
    ws = wb.active
    ws.title = "스팩리스트"

    # 헤더
    for c, (_, label, w) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22

    # 데이터
    spacs = db.get_all_spacs(include_delisted=False)
    computed = [calc.compute_row(s) for s in spacs]
    # 청산임박순 정렬
    computed.sort(key=lambda r: (r.get("days_until_liquidation") or 99999))

    for r_idx, row in enumerate(computed, start=2):
        for c_idx, (key, _, _) in enumerate(HEADERS, start=1):
            val = row.get(key)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER

            # 숫자 포맷
            if key in ("offering_price", "current_price", "liquidation_value"):
                cell.number_format = "#,##0.0"
            elif key in ("rate_y1", "rate_y2", "rate_y3", "total_return_pct", "annualized_return_pct"):
                cell.number_format = "0.00"
                if key == "annualized_return_pct" and isinstance(val, (int, float)):
                    cell.font = Font(bold=True, color="C62828" if val < 0 else "1B5E20")
            elif key == "days_until_liquidation" and isinstance(val, int):
                cell.alignment = Alignment(horizontal="center")
            elif key in ("payment_date", "estimated_liquidation_date"):
                cell.alignment = Alignment(horizontal="center")

            # 행 강조: 청산까지 6개월 이내 → 노랑 / 3개월 이내 → 빨강
            days = row.get("days_until_liquidation")
            if isinstance(days, int):
                if days <= 90:
                    cell.fill = RED_FILL
                elif days <= 180:
                    cell.fill = YELLOW_FILL

            # KSFC 출처면 이율 셀 회색
            if key in ("rate_y1", "rate_y2", "rate_y3") and row.get("rate_source") in ("KSFC", "MIXED"):
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", None):
                    cell.fill = GREY_FILL

    # 틀 고정
    ws.freeze_panes = "C2"

    # ---------- 2. 안내 시트 ----------
    info = wb.create_sheet("안내")
    info["A1"] = "스팩 청산가치 트래커 - 출력본"
    info["A1"].font = Font(bold=True, size=14)
    notes = [
        f"내보낸 시각: {date.today().isoformat()}",
        "",
        "[색상 의미]",
        "  빨강: 청산까지 90일 이하 (관리종목 임박)",
        "  노랑: 청산까지 180일 이하",
        "  회색(이율 셀): KSFC 12개월 거치금 금리 기반 보수적 추정치",
        "",
        "[이율 출처]",
        "  DART  : 전자공시에서 자동 추출한 확정 이율",
        "  KSFC  : 한국증권금융 거치금 금리 기반 추정",
        "  MIXED : 일부 연차는 공시, 일부 연차는 추정",
        "",
        "[청산일 가정]",
        "  납입기일 + 33개월 (2년 9개월) — 통상 청산 시점에 맞춘 보수적 가정",
    ]
    for i, line in enumerate(notes, start=3):
        info.cell(row=i, column=1, value=line)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return str(out_path)


if __name__ == "__main__":
    db.init_db()
    p = export_to_xlsx(Path(__file__).parent / "스팩정리_export.xlsx")
    print(f"saved: {p}")
